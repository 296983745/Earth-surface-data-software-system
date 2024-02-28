"""
#省份级别的拓扑关系
"""

import openpyxl
import pandas as pd
import openpyxl
from content.content_yuyi.simi_hownet import get_mysql_list
import re
import numpy as np
data = openpyxl.load_workbook('E:\\01python_item\pratice\Repetition\Space\province.xlsx')
sheet = data['省拓扑']
def get_relation(privince1,province2):
    # 遍历工作表中的每一行
    # 查找两个元素所在的一行
    if privince1==None or province2==None:
        return 0;
    for row in sheet.iter_rows():
        if row[5].value == privince1 and row[6].value ==province2:
            break
    if row[7].value == "相等":
        return "相等-1"
    elif row[7].value == "相离":
        val=(0.333/32)*(32-row[4].value)
        return "相离-"+str(format(val, '.3f'))
    elif row[7].value == "相接":
        val=0.333+(0.167/32)*(32-row[4].value)
        return "相接-"+str(format(val, '.3f'))

def wordsim(province_names):
    semspace =[]
    for i in range(len(province_names)):
        line = []
        for j in range(len(province_names)):
            simi = get_relation(province_names[i], province_names[j])
            line.append(simi)
        semspace.append(line)
    return semspace


x = get_mysql_list('earthdata', 'data_copy1', 'SpacePlc')
print(x)

# 省份名称字典（包含标准化后的名称）
province_dict = {
    "北京市": "北京",
    "天津市": "天津",
    "河北省": "河北",
    "山西省": "山西",
    "内蒙古自治区": "内蒙古",
    "辽宁省": "辽宁",
    "吉林省": "吉林",
    "黑龙江省": "黑龙江",
    "上海市": "上海",
    "江苏省": "江苏",
    "浙江省": "浙江",
    "安徽省": "安徽",
    "福建省": "福建",
    "江西省": "江西",
    "山东省": "山东",
    "河南省": "河南",
    "湖北省": "湖北",
    "湖南省": "湖南",
    "广东省": "广东",
    "广西壮族自治区": "广西",
    "海南省": "海南",
    "重庆市": "重庆",
    "四川省": "四川",
    "贵州省": "贵州",
    "云南省": "云南",
    "西藏自治区": "西藏",
    "陕西省": "陕西",
    "甘肃省": "甘肃",
    "青海省": "青海",
    "宁夏回族自治区": "宁夏",
    "新疆维吾尔自治区": "新疆",
    "台湾省": "台湾",
    "香港特别行政区": "香港",
    "澳门特别行政区": "澳门",

}
province_dict = {v: k for k, v in province_dict.items()}
province_names= []
# 使用正则表达式匹配省份名称
for text in x:
    pattern = re.compile('|'.join(province_dict.keys()))
    province_name= pattern.findall(str(text))
    # 去重并标准化
    province_name = list(set(province_name))
    province_name = [province_dict[name] for name in province_name]
    province_names.append(province_name)
# 要提取的文本
print(province_names)
province_names = [item for sublist in province_names for item in sublist]
#已经得到省份
print(province_names)
semspace = wordsim(province_names)

print(semspace)
# data_df = pd.DataFrame(semspace)
# # 将文件写入excel表格中
# writer = pd.ExcelWriter('timesimi.xlsx')  # 关键2，创建excel表格
# data_df.to_excel(writer, 'page_1')  # 关键3，float_format 控制精度，将data_df写到hhh表格的第一页中。若多个文件，可以在page_2中写入
# writer.save()  # 关键4

wb = openpyxl.Workbook()

# 获取当前选中的工作表
sheet = wb.active

# 自定义列头
x = get_mysql_list('earthdata', 'data_copy1', 'DataName')
col_headers = [item for sublist in x for item in sublist]

# 自定义行头
row_headers = [item for sublist in x for item in sublist]

# 自定义数据
#semspace
# 写入列头
for col_idx, col_header in enumerate(col_headers, start=2):
    sheet.cell(row=1, column=col_idx, value=col_header)

# 写入行头和数据
for row_idx, row_header in enumerate(row_headers, start=2):
    sheet.cell(row=row_idx, column=1, value=row_header)
    for col_idx, cell_value in enumerate(semspace[row_idx-2], start=2):
        sheet.cell(row=row_idx, column=col_idx, value=cell_value)

# 保存 Excel 文件
wb.save('spacesimisehngshiqu.xlsx')
