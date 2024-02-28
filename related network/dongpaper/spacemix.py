#!/usr/bin/env python
# -*- coding: UTF-8 -*-
'''=================================================
@Project -> File   ：Repetition -> spacemix
@IDE    ：PyCharm
@Author ：haomengqi
@Date   ：2023/12/1 15:10
@Desc   ：
=================================================='''

import time

import pandas as pd
from shapely.geometry import Polygon
from shapely.ops import unary_union
from shapely import affinity
from openpyxl import Workbook

from Time.TimeSimi import TimeRela, reformat
from content.content_yuyi.simi_hownet import get_mysql_list
import threading
import openpyxl
from content.content_yuyi.simi_hownet import get_mysql_list
import re
import pymysql
from content.content_yuyi.simi_yuxian import semDegree
import re



data = openpyxl.load_workbook('D:\\01study\\01python_item\pratice\Repetition\GIS\province+area+river.xlsx')
sheet = data['分省']
# 省份名称字典（包含标准化后的名称）
province_dict = {
    "北京": "北京市",
    "天津": "天津市",
    "河北": "河北省",
    "冀": "河北省",
    "山西": "山西省",
    "内蒙古": "内蒙古自治区",
    "辽宁": "辽宁省",
    "吉林": "吉林省",
    "黑龙江": "黑龙江省",
    "上海": "上海市",
    "长江三角洲": "上海市",
    "江苏": "江苏省",
    "浙江": "浙江省",
    "安徽": "安徽省",
    "福建": "福建省",
    "江西": "江西省",
    "山东": "山东省",
    "河南": "河南省",
    "湖北": "湖北省",
    "长江中游": "湖北省",
    "湖南": "湖南省",
    "广东": "广东省",
    "广西": "广西壮族自治区",
    "海南": "海南省",
    "重庆": "重庆市",
    "四川": "四川省",
    "贵州": "贵州省",
    "云南": "云南省",
    "西藏": "西藏自治区",
    "陕西": "陕西省",
    "甘肃": "甘肃省",
    "青海": "青海省",
    "黄河源": "青海省",
    "宁夏": "宁夏回族自治区",
    "新疆": "新疆维吾尔自治区",
    "台湾": "台湾省",
    "香港": "香港特别行政区",
    "澳门": "澳门特别行政区",
    "辽河": "辽河流域",
    "黄河": "黄河流域",
    "淮河": "淮河流域",
    "珠江": "珠江流域",
    "长江": "长江流域",
    "华北": "华北地区",
    "东北": "东北地区",
    "东三省": "东北地区",
    "华东": "华东地区",
    "中南": "中南地区",
    "西南": "西南地区",
    "西北": "西北地区",
    "成渝": "成渝双城经济圈",
    "黄土高原": "黄土高原",
    "青藏高原": "青藏高原",
    "中国": "中国",
    "全国": "中国",
}
# 判断空间关联度
def simicoordinate(coords1, coords2):
    if coords1 == [] or coords2 == []:
        return 0, "0"
    # 将坐标点转换为 shapely.geometry.Polygon 对象
    poly1 = Polygon(coords1)
    poly2 = Polygon(coords2)
    relation = None
    similarity = 0
    # 判断两个地理区域之间的拓扑关系
    if poly1.equals(poly2):
        similarity = 1.0
        relation = "equals"
    elif poly1.contains(poly2):
        # 计算被包含者占包含者的比例
        ratio = poly2.area / poly1.area
        similarity = ratio * 0.333 + 0.6
        relation = "contains"  # 1包含2 0.5 最多加0.167 包含要边界不相接
    elif poly2.contains(poly1):
        # 交换两个地理区域的顺序，保证第一个地理区域是包含者
        poly1, poly2 = poly2, poly1
        # 计算被包含者占包含者的比例
        ratio = poly2.area / poly1.area
        similarity = ratio * 0.333 + 0.6
        relation = "inside"  # 被包含
    elif poly1.intersects(poly2):
        # 计算两个地理区域的重叠面积和总面积
        intersection = poly1.intersection(poly2)
        union = unary_union([poly1, poly2])
        # 计算重叠面积占总面积的比例
        overlap_ratio = intersection.area / union.area
        similarity = overlap_ratio * 0.167 + 0.4  # 重叠基础相似度是0.5，根据重叠程度加，最多加0.167
        relation = "overlap"
    elif poly1.disjoint(poly2):
        # 计算两个地理区域的距离
        distance = poly1.distance(poly2)
        # 计算相似度，距离越远，相似度越低
        similarity = 1 / (1 + distance) * 0.333  # 相离 最多0.333
        relation = "disjoint"
    return similarity, relation

# 获取四至点
# 5678
def get_zuobiao(privince1):
    # 遍历工作表中的每一行
    # 查找两个元素所在的一行
    found = False
    for row in sheet.iter_rows():
        if row[2].value == privince1:
            found = True
            break
    if found:
        zuobiao = [(31.23, 112.42), (35.05, 112.42), (35.05, 116.12), (31.23, 116.12)]
        minx = row[5].value
        maxx = row[6].value
        miny = row[7].value
        maxy = row[8].value
        zuobiao = [(miny, minx), (maxy, minx), (maxy, maxx), (miny, maxx)]
    else:
        zuobiao = []
    return zuobiao


def getzuobiaos(spacePlc):
    zuobiaos = []
    for i in range(len(spacePlc)):
        zuobiao = get_zuobiao(spacePlc[i])
        zuobiaos.append(zuobiao)
    return zuobiaos


def flatten(lst):
    result = []
    for item in lst:
        if isinstance(item, list):
            result.extend(flatten(item))
        else:
            result.append(item)
    return result

if __name__ == '__main__':
    # ************************************空间************************************
    # 读取位置信息

    spacePlc = get_mysql_list('earthdata', 'datatopic', 'SpacePlc')
    print(spacePlc)
    # print(len(spacePlc))
    province_names = []
    for i, text in enumerate(spacePlc):
        pattern = re.compile('|'.join(province_dict.keys()))
        province_name = pattern.findall(str(text))
        # 去重并标准化
        province_name = list(set(province_name))
        # 将省份名称替换为全称
        for name in province_name:
            if name in province_dict:
                # text = text.replace(name, province_dict[name])
                text = province_dict[name]
        spacePlc[i] = text
    # print(spacePlc)
    spacePlc = flatten(spacePlc)
    # print(spacePlc)
    # print(len(spacePlc))
    zuobiaoschn = getzuobiaos(spacePlc)
    # 得到用坐标表示的位置信息，没有坐标的为空
    # print(zuobiaoschn)
    # 从数据库中取元素
    Southernmost_Latitude = get_mysql_list('earthdata', 'engnasa', 'Southernmost_Latitude')  # 南边
    Northernmost_Latitude = get_mysql_list('earthdata', 'engnasa', 'Northernmost_Latitude')  # 北边
    Westernmost_Longitude = get_mysql_list('earthdata', 'engnasa', 'Westernmost_Longitude')  # 西边
    Easternmost_Longitude = get_mysql_list('earthdata', 'engnasa', 'Easternmost_Longitude')  # 东边

    zuobiaoseng = []
    for i in range(len(Southernmost_Latitude)):
        if Westernmost_Longitude[i][0] == '':
            Westernmost_Longitude[i][0] = 0
        if Southernmost_Latitude[i][0] == '':
            Southernmost_Latitude[i][0] = 0
        if Northernmost_Latitude[i][0] == '':
            Northernmost_Latitude[i][0] = 0
        if Easternmost_Longitude[i][0] == '':
            Easternmost_Longitude[i][0] = 0
        Westernmost = float(Westernmost_Longitude[i][0])
        Southernmost = float(Southernmost_Latitude[i][0])
        Northernmost = float(Northernmost_Latitude[i][0])
        Easternmost = float(Easternmost_Longitude[i][0])
        zuobiao = [(Westernmost, Southernmost), (Westernmost, Northernmost), (Easternmost, Northernmost),
                   (Easternmost, Southernmost)]
        zuobiaoseng.append(zuobiao)
    # print(zuobiaoseng)
    zuobiaos=zuobiaoseng+zuobiaoschn
    print(len(zuobiaos))
    workbook = Workbook()
    sheet = workbook.active
    column_names = ["数据集1", "数据集2", "关键词1", "关键词2", "空间关联度"]
    # 写入列名
    namechn= get_mysql_list('earthdata', 'datatopic', 'DataName')
    nameeng= get_mysql_list('earthdata', 'engnasa', 'Entry_Title')
    name=nameeng+namechn
    print(len(name))
    sheet.append(column_names)
    eg=[(-73.48442, 41.26774), (-73.48442, 42.84917), (-69.95562, 42.84917), (-69.95562, 41.26774)]
    for i in range(len(zuobiaos)):
        # for j in range(i+1,len(zuobiaos)):
        simis, simispace = simicoordinate(eg, zuobiaos[i])
        print(simis)
        print(simispace)
        simi = str(format(simis, '.3f')) + "-" + simispace
        # data = [f"{name[i]}", f"{name[j]}", f"{zuobiaos[i]}", f"{zuobiaos[j]}",f"{simispace+str(format(simis, '.3f'))}"]
        # print(name[i][0])
        data = ["Notes and observations made of Antarctic Petrels and Silver Grey Fulmars in the Windmill Islands, 1972", name[i][0], str(eg), str(zuobiaos[i]), simi]
        sheet.append(data)
    workbook.save("mixspace1.xlsx")